# ⚠️ Compliance Notice:
# Assistive mode only. Do NOT overwrite validated cells/ranges or macros in `MTCR Data.xlsm`.
# Always create a NEW worksheet for AI outputs, prefixed with "AI_".
# Never rename, delete, or edit existing validated sheets.
# Keep a timestamped backup copy before any write.
# Log every write action (JSONL) for QA traceability.

"""
MTCR Excel Writer - Module 3 (M3)

Safe Excel writer that creates new worksheets with AI_ prefixed columns.
Never modifies validated sheets or macros.

Created by: Navid Broumandfar (Service Analytics, CHP, bioMérieux)
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
import shutil
from typing import List, Optional

import pandas as pd

# ---- Config defaults
DEFAULT_WORKBOOK = "/mnt/data/MTCR Data.xlsm"
DEFAULT_SHEET_IN = "Quality Review"
DEFAULT_SHEET_OUT_PREFIX = "AI_Quality_Review"
LOG_PATH = "logs/mtcr_writer.jsonl"

REQUIRED_AI_COLS = [
    "AI_ReasonSuggestion",
    "AI_Confidence",
    "AI_CommentStandardized",
    "AI_RationaleShort",
    "AI_ModelVersion",
]


def timestamp(fmt="%Y%m%d_%H%M%S") -> str:
    return dt.datetime.now().strftime(fmt)


def backup_workbook(path: str) -> str:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Workbook not found: {path}")
    root, ext = os.path.splitext(path)
    bkp = f"{root}.backup_{timestamp()}{ext}"
    os.makedirs(os.path.dirname(bkp), exist_ok=True)
    shutil.copy2(path, bkp)
    return bkp


def log_write(event: dict):
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
    event = {"ts": dt.datetime.now().isoformat(), **event}
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")


def load_quality_review(
    from_m1: bool, sheet_in: str, input_csv: Optional[str]
) -> pd.DataFrame:
    if from_m1:
        # Lazy import to avoid hard dependency when using CSV mode
        try:
            from src.excel.mtcr_reader import read_quality_review  # M1 entrypoint
            from src.utils.config_loader import load_config
        except Exception as e:
            raise RuntimeError("Failed to import/read via M1 (mtcr_reader).") from e

        # Load config and read data
        config = load_config()
        df, _ = read_quality_review(config)
        return df
    else:
        if not input_csv:
            raise ValueError("--input_csv must be provided when --from_m1 is false")
        return pd.read_csv(input_csv)


def ensure_ai_columns(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in REQUIRED_AI_COLS if c not in df.columns]
    if missing:
        raise ValueError(
            "Missing required AI_ columns: "
            + ", ".join(missing)
            + ". Ensure M2 outputs are present."
        )
    return df


def compose_output_df(df: pd.DataFrame, source_cols: List[str]) -> pd.DataFrame:
    keep = []
    for c in source_cols:
        if c not in df.columns:
            raise ValueError(f"Requested source column not in DataFrame: {c}")
        keep.append(c)
    keep.extend(REQUIRED_AI_COLS)
    return df.loc[:, keep].copy()


def _next_available_sheet_name(existing: List[str], base: str) -> str:
    # base should already include date: e.g., AI_Quality_Review_20251028
    if base not in existing:
        return base
    i = 2
    while True:
        cand = f"{base}_{i}"
        if cand not in existing:
            return cand
        i += 1


def write_new_sheet_win32(workbook_path: str, sheet_name: str, df: pd.DataFrame):
    import pythoncom  # type: ignore
    import win32com.client as win32  # type: ignore

    pythoncom.CoInitialize()
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(workbook_path)
        # Collect existing names to avoid collisions
        existing_names = [s.Name for s in wb.Worksheets]
        sheet_name_final = _next_available_sheet_name(existing_names, sheet_name)
        ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
        ws.Name = sheet_name_final

        # Fast write: headers
        headers = list(df.columns)
        ws.Range(ws.Cells(1, 1), ws.Cells(1, len(headers))).Value = [headers]

        # Values as 2D array
        if len(df) > 0:
            data = df.values.tolist()
            ws.Range(ws.Cells(2, 1), ws.Cells(len(data) + 1, len(headers))).Value = data

        wb.Save()
        new_sheet = sheet_name_final
    finally:
        wb.Close(SaveChanges=True)
        excel.Quit()
        pythoncom.CoUninitialize()
    return new_sheet


def write_new_sheet_openpyxl(workbook_path: str, sheet_name: str, df: pd.DataFrame):
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, keep_vba=True)
    existing = wb.sheetnames
    sheet_name_final = _next_available_sheet_name(existing, sheet_name)
    ws = wb.create_sheet(sheet_name_final)

    # Headers
    ws.append(list(df.columns))
    # Rows
    for _, row in df.iterrows():
        ws.append(row.tolist())

    wb.save(workbook_path)
    return sheet_name_final


def run_writer(
    workbook: str,
    sheet_in: str,
    sheet_out_prefix: str,
    source_cols_csv: str,
    from_m1: bool,
    input_csv: Optional[str],
    include_date: bool = True,
):
    # 1) backup
    backup_path = backup_workbook(workbook)

    # 2) load input
    df = load_quality_review(from_m1=from_m1, sheet_in=sheet_in, input_csv=input_csv)

    # 3) ensure AI cols
    df = ensure_ai_columns(df)

    # 4) compose output
    source_cols = [c.strip() for c in (source_cols_csv or "").split(",") if c.strip()]
    out_df = compose_output_df(df, source_cols)

    # 5) sheet name
    date_token = (
        dt.datetime.now().strftime("%Y%m%d") if include_date else timestamp("%Y%m%d")
    )
    target_sheet = f"{sheet_out_prefix}_{date_token}"

    # 6) write via COM if available, else fallback
    wrote_sheet = None
    mode = "csv" if not from_m1 else "pipeline"
    try:
        import win32com.client  # noqa: F401

        wrote_sheet = write_new_sheet_win32(workbook, target_sheet, out_df)
        engine = "win32com"
    except Exception:
        # Fallback to openpyxl
        wrote_sheet = write_new_sheet_openpyxl(workbook, target_sheet, out_df)
        engine = "openpyxl"

    # 7) log
    event = {
        "workbook": workbook,
        "new_sheet": wrote_sheet,
        "rows": int(len(out_df)),
        "columns": list(out_df.columns),
        "backup_path": backup_path,
        "engine": engine,
        "mode": mode,
    }
    log_write(event)
    return event


def main(argv=None):
    p = argparse.ArgumentParser(
        description="Safe Excel writer for MTCR AI outputs (M3)."
    )
    p.add_argument("--workbook", default=DEFAULT_WORKBOOK)
    p.add_argument("--sheet_in", default=DEFAULT_SHEET_IN)
    p.add_argument("--sheet_out_prefix", default=DEFAULT_SHEET_OUT_PREFIX)
    p.add_argument("--source_cols", default="CaseID,Product,ErrorCode,ReviewComment")
    p.add_argument("--from_m1", type=lambda s: s.lower() == "true", default=False)
    p.add_argument("--input_csv", default=None)
    args = p.parse_args(argv)

    # Optional: pipeline mode invokes M2 automatically to generate AI_ columns if they're missing.
    # For now, we assume input has AI_ columns (either from M2 or a prepared CSV).
    # If you want full pipeline here, import and run your M2 entrypoint before ensure_ai_columns().

    event = run_writer(
        workbook=args.workbook,
        sheet_in=args.sheet_in,
        sheet_out_prefix=args.sheet_out_prefix,
        source_cols_csv=args.source_cols,
        from_m1=args.from_m1,
        input_csv=args.input_csv,
    )
    print(json.dumps(event, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    sys.exit(main())
