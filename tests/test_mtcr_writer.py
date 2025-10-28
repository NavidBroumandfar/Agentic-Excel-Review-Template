# ⚠️ Compliance Notice:
# This test uses a temporary .xlsx to avoid touching the macro workbook.

import os
import tempfile
import pandas as pd
from unittest.mock import patch, MagicMock
from pathlib import Path

# pytest is optional for testing
try:
    import pytest
except ImportError:
    pytest = None

from src.excel.mtcr_writer import (
    write_new_sheet_openpyxl,
    compose_output_df,
    ensure_ai_columns,
    backup_workbook,
    log_write,
    _next_available_sheet_name,
    REQUIRED_AI_COLS,
)


def _dummy_df():
    base = pd.DataFrame(
        {
            "CaseID": [1, 2],
            "Product": ["X", "Y"],
            "ErrorCode": ["E1", "E2"],
            "ReviewComment": ["foo", "bar"],
            "AI_ReasonSuggestion": ["Reason A", "Reason B"],
            "AI_Confidence": [0.88, 0.73],
            "AI_CommentStandardized": ["std A", "std B"],
            "AI_RationaleShort": ["why A", "why B"],
            "AI_ModelVersion": ["m2-v1", "m2-v1"],
        }
    )
    return base


def test_openpyxl_write_new_sheet(tmp_path=None):
    if tmp_path is None:
        import tempfile

        tmp_path = tempfile.mkdtemp()
    # Create a temporary workbook (*.xlsx is fine for openpyxl path)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Dummy"
    xlsx_path = os.path.join(tmp_path, "temp.xlsx")
    wb.save(xlsx_path)

    df = _dummy_df()
    # Validate AI columns present
    ensure_ai_columns(df)
    # Compose output: select two source cols + AI cols
    out_df = compose_output_df(df, ["CaseID", "Product"])
    assert out_df.shape[1] == 2 + len(REQUIRED_AI_COLS)

    # Write new AI sheet
    new_sheet = write_new_sheet_openpyxl(
        xlsx_path, "AI_Quality_Review_20991231", out_df
    )
    assert new_sheet.startswith("AI_Quality_Review_20991231")

    from openpyxl import load_workbook

    wb2 = load_workbook(xlsx_path)
    assert new_sheet in wb2.sheetnames

    ws2 = wb2[new_sheet]
    # Header row
    headers = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
    assert headers == list(out_df.columns)

    # Two data rows + header
    rows = list(ws2.iter_rows(values_only=True))
    assert len(rows) == 1 + len(out_df)


def test_next_available_sheet_name():
    # Test basic case
    existing = ["Sheet1", "Sheet2"]
    assert _next_available_sheet_name(existing, "AI_Test") == "AI_Test"

    # Test collision case
    existing = ["Sheet1", "AI_Test", "Sheet2"]
    assert _next_available_sheet_name(existing, "AI_Test") == "AI_Test_2"

    # Test multiple collisions
    existing = ["Sheet1", "AI_Test", "AI_Test_2", "Sheet2"]
    assert _next_available_sheet_name(existing, "AI_Test") == "AI_Test_3"


def test_ensure_ai_columns():
    df = _dummy_df()
    # Should pass with all required columns
    result = ensure_ai_columns(df)
    assert result is df  # Should return same object

    # Should fail with missing columns
    df_missing = df.drop(columns=["AI_Confidence"])
    try:
        ensure_ai_columns(df_missing)
        assert False, "Should have raised ValueError"
    except ValueError as e:
        assert "Missing required AI_ columns" in str(e)


def test_compose_output_df():
    df = _dummy_df()

    # Test with valid source columns
    result = compose_output_df(df, ["CaseID", "Product"])
    expected_cols = ["CaseID", "Product"] + REQUIRED_AI_COLS
    assert list(result.columns) == expected_cols
    assert len(result) == 2

    # Test with missing source column
    try:
        compose_output_df(df, ["CaseID", "NonExistentColumn"])
        assert False, "Should have raised ValueError"
    except ValueError as e:
        assert "Requested source column not in DataFrame" in str(e)


def test_backup_workbook(tmp_path=None):
    if tmp_path is None:
        import tempfile

        tmp_path = tempfile.mkdtemp()
    # Create a test file
    test_file = tmp_path / "test.xlsx"
    test_file.write_text("test content")

    # Test backup creation
    backup_path = backup_workbook(str(test_file))
    assert backup_path != str(test_file)
    assert os.path.exists(backup_path)
    assert os.path.exists(str(test_file))  # Original should still exist

    # Test backup content matches original
    with open(str(test_file), "r") as f:
        original_content = f.read()
    with open(backup_path, "r") as f:
        backup_content = f.read()
    assert original_content == backup_content

    # Test file not found
    try:
        backup_workbook("nonexistent.xlsx")
        assert False, "Should have raised FileNotFoundError"
    except FileNotFoundError:
        pass  # Expected


def test_log_write(tmp_path=None):
    if tmp_path is None:
        import tempfile

        tmp_path = tempfile.mkdtemp()
    # Mock the LOG_PATH to use temp directory
    with patch("src.excel.mtcr_writer.LOG_PATH", str(tmp_path / "test_log.jsonl")):
        # Test log write
        test_event = {"test": "data", "number": 42}
        log_write(test_event)

        # Verify log file was created and contains the event
        log_file = tmp_path / "test_log.jsonl"
        assert log_file.exists()

        with open(log_file, "r") as f:
            log_content = f.read().strip()

        # Should be a single JSON line
        import json

        logged_event = json.loads(log_content)
        assert logged_event["test"] == "data"
        assert logged_event["number"] == 42
        assert "ts" in logged_event  # Should have timestamp


def test_run_writer_csv_mode(tmp_path=None):
    """Test the run_writer function in CSV mode"""
    if tmp_path is None:
        import tempfile

        tmp_path = tempfile.mkdtemp()
    from src.excel.mtcr_writer import run_writer

    # Create test CSV with AI columns
    test_csv = tmp_path / "test_input.csv"
    df = _dummy_df()
    df.to_csv(test_csv, index=False)

    # Create test workbook
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Review"
    test_workbook = tmp_path / "test_workbook.xlsx"
    wb.save(test_workbook)

    # Mock backup_workbook to avoid actual file operations
    with patch("src.excel.mtcr_writer.backup_workbook") as mock_backup:
        mock_backup.return_value = str(tmp_path / "backup.xlsx")

        # Mock log_write to avoid file operations
        with patch("src.excel.mtcr_writer.log_write") as mock_log:
            # Run writer
            result = run_writer(
                workbook=str(test_workbook),
                sheet_in="Quality Review",
                sheet_out_prefix="AI_Quality_Review",
                source_cols_csv="CaseID,Product",
                from_m1=False,
                input_csv=str(test_csv),
            )

            # Verify result structure
            assert "workbook" in result
            assert "new_sheet" in result
            assert "rows" in result
            assert "columns" in result
            assert "backup_path" in result
            assert "engine" in result
            assert "mode" in result
            assert result["mode"] == "csv"

            # Verify backup was called
            mock_backup.assert_called_once()

            # Verify log was called
            mock_log.assert_called_once()


def test_run_writer_missing_ai_columns(tmp_path=None):
    """Test run_writer fails gracefully when AI columns are missing"""
    if tmp_path is None:
        import tempfile

        tmp_path = tempfile.mkdtemp()
    from src.excel.mtcr_writer import run_writer

    # Create test CSV without AI columns
    test_csv = tmp_path / "test_input.csv"
    df = pd.DataFrame(
        {
            "CaseID": [1, 2],
            "Product": ["X", "Y"],
            "ReviewComment": ["foo", "bar"],
        }
    )
    df.to_csv(test_csv, index=False)

    # Create test workbook
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Review"
    test_workbook = tmp_path / "test_workbook.xlsx"
    wb.save(test_workbook)

    # Should raise ValueError for missing AI columns
    try:
        run_writer(
            workbook=str(test_workbook),
            sheet_in="Quality Review",
            sheet_out_prefix="AI_Quality_Review",
            source_cols_csv="CaseID,Product",
            from_m1=False,
            input_csv=str(test_csv),
        )
        assert False, "Should have raised ValueError"
    except ValueError as e:
        assert "Missing required AI_ columns" in str(e)


def test_run_writer_m1_mode():
    """Test run_writer in M1 mode (requires mocking M1 components)"""
    from src.excel.mtcr_writer import run_writer

    # Mock the M1 components
    mock_df = _dummy_df()

    with patch("src.excel.mtcr_writer.load_quality_review") as mock_load:
        mock_load.return_value = mock_df

        with patch("src.excel.mtcr_writer.backup_workbook") as mock_backup:
            mock_backup.return_value = "/tmp/backup.xlsx"

            with patch("src.excel.mtcr_writer.write_new_sheet_openpyxl") as mock_write:
                mock_write.return_value = "AI_Quality_Review_20251028"

                with patch("src.excel.mtcr_writer.log_write") as mock_log:
                    # Run writer in M1 mode
                    result = run_writer(
                        workbook="/tmp/test.xlsx",
                        sheet_in="Quality Review",
                        sheet_out_prefix="AI_Quality_Review",
                        source_cols_csv="CaseID,Product",
                        from_m1=True,
                        input_csv=None,
                    )

                    # Verify M1 was called
                    mock_load.assert_called_once_with(
                        from_m1=True, sheet_in="Quality Review", input_csv=None
                    )

                    # Verify result
                    assert result["mode"] == "pipeline"
                    assert "new_sheet" in result


def test_main_function():
    """Test the main CLI function"""
    from src.excel.mtcr_writer import main

    # Test with help
    try:
        main(["--help"])
        assert False, "Should have raised SystemExit"
    except SystemExit:
        pass  # Expected


def test_required_ai_columns_constant():
    """Test that REQUIRED_AI_COLS contains expected columns"""
    expected_cols = [
        "AI_ReasonSuggestion",
        "AI_Confidence",
        "AI_CommentStandardized",
        "AI_RationaleShort",
        "AI_ModelVersion",
    ]
    assert REQUIRED_AI_COLS == expected_cols
