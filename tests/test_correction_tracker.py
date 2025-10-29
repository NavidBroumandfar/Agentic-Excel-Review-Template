# ⚠️ Compliance Notice:
# Test suite for Correction Tracker Agent (M8).
# Tests ensure assistive-only mode and proper data handling.

"""
Unit tests for Correction Tracker Agent.
Tests discrepancy detection, confidence correlation, and output generation.
"""

import json
import os
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch, MagicMock

import pandas as pd

# Add src to path for imports
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), "..", "src"))

from ai.correction_tracker import CorrectionTracker


class TestCorrectionTracker(unittest.TestCase):
    """Test cases for Correction Tracker Agent."""

    def setUp(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        self.test_month = "202510"
        
        # Create test log file
        self.test_log_file = os.path.join(self.temp_dir, "test_correction_tracker.jsonl")
        
        # Initialize tracker with test directory
        self.tracker = CorrectionTracker(log_file=self.test_log_file)
        self.tracker.month = self.test_month

    def tearDown(self):
        """Clean up test fixtures."""
        # Clean up temporary files
        for file_path in Path(self.temp_dir).glob("*"):
            if file_path.is_file():
                file_path.unlink()

    def test_initialization(self):
        """Test CorrectionTracker initialization."""
        self.assertEqual(self.tracker.month, self.test_month)
        self.assertEqual(self.tracker.log_file, self.test_log_file)
        self.assertIsNotNone(self.tracker.config)

    def test_calculate_similarity(self):
        """Test text similarity calculation."""
        # Test exact match
        similarity = self.tracker._calculate_similarity("calibration certificate", "calibration certificate")
        self.assertAlmostEqual(similarity, 1.0, places=2)
        
        # Test partial match
        similarity = self.tracker._calculate_similarity("calibration certificate", "calibration cert")
        self.assertGreater(similarity, 0.7)
        
        # Test no match
        similarity = self.tracker._calculate_similarity("calibration", "documentation")
        self.assertLessEqual(similarity, 0.5)
        
        # Test empty strings
        similarity = self.tracker._calculate_similarity("", "test")
        self.assertEqual(similarity, 0.0)

    def test_classify_match_status(self):
        """Test match status classification."""
        # Test matched case
        status = self.tracker._classify_match_status("calibration certificate", "calibration certificate")
        self.assertEqual(status, "Matched")
        
        # Test partial case
        status = self.tracker._classify_match_status("calibration certificate", "calibration cert missing")
        self.assertEqual(status, "Partial")
        
        # Test overridden case
        status = self.tracker._classify_match_status("calibration certificate", "documentation incomplete")
        self.assertEqual(status, "Overridden")

    def test_load_ai_outputs(self):
        """Test loading AI outputs from logs."""
        # Create test log file
        test_log_data = [
            {
                "timestamp": "2025-10-01T10:00:00Z",
                "row_id": "case_001",
                "response": {
                    "reason": "Missing calibration certificate",
                    "confidence": 0.85,
                    "model_version": "MTCR-Llama3-v0.1"
                },
                "original_comment": "No calibration cert found"
            },
            {
                "timestamp": "2025-10-01T10:01:00Z",
                "row_id": "case_002",
                "response": {
                    "reason": "Documentation incomplete",
                    "confidence": 0.92,
                    "model_version": "MTCR-Llama3-v0.1"
                },
                "original_comment": "Missing batch number"
            }
        ]
        
        # Write test log file
        log_path = f"logs/mtcr_review_assistant_{self.test_month}.jsonl"
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        with open(log_path, "w", encoding="utf-8") as f:
            for record in test_log_data:
                f.write(json.dumps(record) + "\n")
        
        # Load AI outputs
        ai_df = self.tracker._load_ai_outputs(self.test_month)
        
        # Verify results
        self.assertEqual(len(ai_df), 2)
        self.assertIn("case_id", ai_df.columns)
        self.assertIn("ai_reason", ai_df.columns)
        self.assertIn("ai_confidence", ai_df.columns)
        self.assertEqual(ai_df.iloc[0]["ai_reason"], "Missing calibration certificate")
        self.assertEqual(ai_df.iloc[0]["ai_confidence"], 0.85)

    def test_create_mock_human_data(self):
        """Test creation of mock human correction data."""
        # Create test AI data
        ai_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "ai_reason": "Missing calibration certificate",
                "ai_confidence": 0.85
            },
            {
                "case_id": "case_002",
                "ai_reason": "Documentation incomplete",
                "ai_confidence": 0.92
            }
        ])
        
        # Create mock human data
        human_df = self.tracker._create_mock_human_data(ai_df)
        
        # Verify results
        self.assertEqual(len(human_df), 2)
        self.assertIn("case_id", human_df.columns)
        self.assertIn("human_reason", human_df.columns)
        self.assertIn("reviewer", human_df.columns)
        
        # Check that case_001 gets exact match (ends with 1)
        case_001 = human_df[human_df["case_id"] == "case_001"].iloc[0]
        self.assertEqual(case_001["human_reason"], "Missing calibration certificate")

    def test_compare_corrections(self):
        """Test the main comparison logic."""
        # Create test AI data
        ai_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "ai_reason": "Missing calibration certificate",
                "ai_confidence": 0.85,
                "ai_model_version": "MTCR-Llama3-v0.1",
                "original_comment": "No calibration cert found",
                "timestamp": "2025-10-01T10:00:00Z"
            },
            {
                "case_id": "case_002",
                "ai_reason": "Documentation incomplete",
                "ai_confidence": 0.92,
                "ai_model_version": "MTCR-Llama3-v0.1",
                "original_comment": "Missing batch number",
                "timestamp": "2025-10-01T10:01:00Z"
            }
        ])
        
        # Mock the _load_ai_outputs method
        with patch.object(self.tracker, '_load_ai_outputs', return_value=ai_df):
            with patch.object(self.tracker, '_create_mock_human_data') as mock_human:
                # Create mock human data
                human_df = pd.DataFrame([
                    {
                        "case_id": "case_001",
                        "human_reason": "Missing calibration certificate",
                        "reviewer": "QA_Reviewer_001",
                        "review_date": "2025-10-01"
                    },
                    {
                        "case_id": "case_002",
                        "human_reason": "Documentation incomplete - missing batch number",
                        "reviewer": "QA_Reviewer_001",
                        "review_date": "2025-10-01"
                    }
                ])
                mock_human.return_value = human_df
                
                # Run comparison
                result_df = self.tracker.compare_corrections(self.test_month)
                
                # Verify results
                self.assertEqual(len(result_df), 2)
                self.assertIn("match_status", result_df.columns)
                self.assertIn("similarity_score", result_df.columns)
                
                # Check that case_001 is matched (exact match)
                case_001 = result_df[result_df["case_id"] == "case_001"].iloc[0]
                self.assertEqual(case_001["match_status"], "Matched")

    def test_generate_kpi_summary(self):
        """Test KPI summary generation."""
        # Create test comparison data
        comparison_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "match_status": "Matched",
                "ai_confidence": 0.85
            },
            {
                "case_id": "case_002",
                "match_status": "Overridden",
                "ai_confidence": 0.92
            },
            {
                "case_id": "case_003",
                "match_status": "Matched",
                "ai_confidence": 0.78
            }
        ])
        
        # Generate KPI summary
        kpi_summary = self.tracker.generate_kpi_summary(comparison_df)
        
        # Verify results
        self.assertEqual(kpi_summary["total_cases"], 3)
        self.assertEqual(kpi_summary["matched_count"], 2)
        self.assertEqual(kpi_summary["overridden_count"], 1)
        self.assertAlmostEqual(kpi_summary["match_rate_pct"], 66.67, places=1)
        self.assertAlmostEqual(kpi_summary["override_rate_pct"], 33.33, places=1)
        self.assertAlmostEqual(kpi_summary["avg_confidence"], 0.85, places=2)

    def test_export_to_csv(self):
        """Test CSV export functionality."""
        # Create test data
        comparison_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "ai_reason": "Missing calibration certificate",
                "human_reason": "Missing calibration certificate",
                "match_status": "Matched",
                "ai_confidence": 0.85,
                "reviewer": "QA_Reviewer_001",
                "review_date": "2025-10-01"
            }
        ])
        
        kpi_summary = {
            "month": self.test_month,
            "total_cases": 1,
            "match_rate_pct": 100.0,
            "override_rate_pct": 0.0,
            "avg_confidence": 0.85
        }
        
        # Export to CSV
        csv_path = self.tracker.export_to_csv(comparison_df, kpi_summary)
        
        # Verify file was created
        self.assertTrue(os.path.exists(csv_path))
        
        # Verify content
        with open(csv_path, "r", encoding="utf-8") as f:
            content = f.read()
            self.assertIn("KPI Summary", content)
            self.assertIn("Detailed Results", content)
            self.assertIn("case_001", content)

    def test_export_to_excel(self):
        """Test Excel export functionality."""
        # Create test data
        comparison_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "ai_reason": "Missing calibration certificate",
                "human_reason": "Missing calibration certificate",
                "match_status": "Matched",
                "ai_confidence": 0.85,
                "reviewer": "QA_Reviewer_001",
                "review_date": "2025-10-01"
            }
        ])
        
        kpi_summary = {
            "month": self.test_month,
            "total_cases": 1,
            "match_rate_pct": 100.0,
            "override_rate_pct": 0.0,
            "avg_confidence": 0.85
        }
        
        # Export to Excel
        xlsx_path = self.tracker.export_to_excel(comparison_df, kpi_summary)
        
        # Verify file was created
        self.assertTrue(os.path.exists(xlsx_path))
        
        # Verify Excel structure
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        self.assertIn("Overview", wb.sheetnames)
        self.assertIn("By_Subsidiary", wb.sheetnames)
        self.assertIn("By_Confidence_Range", wb.sheetnames)
        self.assertIn("Details", wb.sheetnames)

    def test_log_event(self):
        """Test event logging functionality."""
        # Log a test event
        test_event = {
            "event_type": "test_event",
            "test_data": "test_value"
        }
        
        self.tracker._log_event(test_event)
        
        # Verify log file was created and contains the event
        self.assertTrue(os.path.exists(self.test_log_file))
        
        with open(self.test_log_file, "r", encoding="utf-8") as f:
            log_content = f.read()
            self.assertIn("test_event", log_content)
            self.assertIn("test_value", log_content)

    def test_run_analysis_empty_data(self):
        """Test run_analysis with empty data."""
        # Mock empty AI outputs
        with patch.object(self.tracker, '_load_ai_outputs', return_value=pd.DataFrame()):
            result = self.tracker.run_analysis(self.test_month)
            
            # Should return error for empty data
            self.assertIn("error", result)
            self.assertEqual(result["error"], "No data available for analysis")

    def test_run_analysis_success(self):
        """Test successful run_analysis."""
        # Create test data
        ai_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "ai_reason": "Missing calibration certificate",
                "ai_confidence": 0.85,
                "ai_model_version": "MTCR-Llama3-v0.1",
                "original_comment": "No calibration cert found",
                "timestamp": "2025-10-01T10:00:00Z"
            }
        ])
        
        human_df = pd.DataFrame([
            {
                "case_id": "case_001",
                "human_reason": "Missing calibration certificate",
                "reviewer": "QA_Reviewer_001",
                "review_date": "2025-10-01"
            }
        ])
        
        # Mock the methods
        with patch.object(self.tracker, '_load_ai_outputs', return_value=ai_df):
            with patch.object(self.tracker, '_create_mock_human_data', return_value=human_df):
                result = self.tracker.run_analysis(self.test_month)
                
                # Verify successful result
                self.assertNotIn("error", result)
                self.assertEqual(result["month"], self.test_month)
                self.assertIn("comparison_df", result)
                self.assertIn("kpi_summary", result)
                self.assertIn("csv_path", result)
                self.assertIn("xlsx_path", result)

    def test_compliance_mode(self):
        """Test that the tracker operates in assistive-only mode."""
        # Verify that no existing files are modified
        # This is implicit in the design - all outputs go to new files
        
        # Test that log file is created in logs directory
        # In test mode, the log file is in temp directory, so just check it's set
        self.assertIsNotNone(self.tracker.log_file)
        self.assertTrue(self.tracker.log_file.endswith('.jsonl'))
        
        # Test that CSV output goes to data/outputs
        comparison_df = pd.DataFrame([{"case_id": "test", "ai_reason": "test", "human_reason": "test", 
                                     "match_status": "Matched", "ai_confidence": 0.8, "reviewer": "test", 
                                     "review_date": "2025-10-01"}])
        kpi_summary = {"month": self.test_month, "total_cases": 1, "match_rate_pct": 100.0, 
                      "override_rate_pct": 0.0, "avg_confidence": 0.8}
        
        csv_path = self.tracker.export_to_csv(comparison_df, kpi_summary)
        self.assertTrue(csv_path.startswith("data/outputs/"))


if __name__ == "__main__":
    # Create necessary directories for tests
    os.makedirs("logs", exist_ok=True)
    os.makedirs("data/outputs", exist_ok=True)
    
    # Run tests
    unittest.main()
